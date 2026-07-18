"""导入导出业务逻辑"""

import io
from datetime import datetime

import pandas as pd
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.test_case import TestCase, PriorityEnum, CaseStatusEnum
from app.models.test_plan import TestPlan
from app.models.test_plan_item import TestPlanItem, ExecutionStatusEnum
from app.services import test_plan_service


# ── 导出 ──

HEADER_FILL = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_case_sheet(ws, cases: list[TestCase]):
    """将用例写入工作表"""
    headers = ["ID", "标题", "模块", "优先级", "状态", "前置条件", "步骤", "标签", "创建时间", "更新时间"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for row_idx, case in enumerate(cases, 2):
        values = [
            case.id, case.title, case.module,
            case.priority.value if isinstance(case.priority, PriorityEnum) else case.priority,
            case.status.value if isinstance(case.status, CaseStatusEnum) else case.status,
            case.precondition or "",
            "\n".join([f"{s.get('seq', '')}. {s.get('action', '')} → {s.get('expected', '')}" for s in (case.steps or [])]),
            ", ".join(case.tags) if case.tags else "",
            case.created_at.strftime("%Y-%m-%d %H:%M") if case.created_at else "",
            case.updated_at.strftime("%Y-%m-%d %H:%M") if case.updated_at else "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 设置列宽
    widths = [6, 30, 12, 8, 8, 25, 50, 20, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w


async def export_test_cases_to_excel(db: AsyncSession, filters: dict | None = None) -> io.BytesIO:
    """导出用例为 Excel"""
    query = select(TestCase).order_by(TestCase.module, TestCase.id)
    if filters:
        if filters.get("module"):
            query = query.where(TestCase.module == filters["module"])
        if filters.get("priority"):
            query = query.where(TestCase.priority == filters["priority"])
        if filters.get("status"):
            query = query.where(TestCase.status == filters["status"])

    result = await db.execute(query)
    cases = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    _write_case_sheet(ws, cases)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_test_plan_to_excel(db: AsyncSession, plan_id: int) -> io.BytesIO | None:
    """导出测试计划为 Excel"""
    plan = await test_plan_service.get_test_plan(db, plan_id)
    if not plan:
        return None

    wb = Workbook()
    # 计划信息页
    ws_info = wb.active
    ws_info.title = "计划信息"
    summary = test_plan_service.compute_summary(plan.items)
    info_data = [
        ["计划名称", plan.name],
        ["描述", plan.description or ""],
        ["状态", {"draft": "草稿", "in_progress": "进行中", "completed": "已完成", "archived": "已归档"}.get(plan.status.value if hasattr(plan.status, 'value') else str(plan.status), str(plan.status))],
        ["通过", summary["passed"]],
        ["失败", summary["failed"]],
        ["阻塞", summary["blocked"]],
        ["跳过", summary["skipped"]],
        ["未执行", summary["untested"]],
    ]
    for row, (k, v) in enumerate(info_data, 1):
        ws_info.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws_info.cell(row=row, column=2, value=str(v))
    ws_info.column_dimensions['A'].width = 15
    ws_info.column_dimensions['B'].width = 40

    # 用例执行页
    ws_cases = wb.create_sheet("用例执行结果")
    headers = ["用例ID", "用例标题", "模块", "优先级", "执行状态", "执行人", "实际结果", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws_cases.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for row_idx, item in enumerate(plan.items, 2):
        exe = {
            "pass": "Pass", "fail": "Fail", "nt": "NT", "na": "NA", None: "未执行",
        }
        status_val = item.execution_status.value if item.execution_status else None
        values = [
            item.test_case_id,
            item.test_case.title if item.test_case else "",
            item.test_case.module if item.test_case else "",
            item.test_case.priority.value if item.test_case and item.test_case.priority else "",
            exe.get(status_val, "未执行"),
            item.executed_by or "",
            item.actual_result or "",
            item.notes or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws_cases.cell(row=row_idx, column=col, value=val)
            cell.border = THIN_BORDER

    widths = [8, 30, 12, 8, 10, 10, 30, 25]
    for col, w in enumerate(widths, 1):
        ws_cases.column_dimensions[ws_cases.cell(row=1, column=col).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── 导入 ──

# 测试用例列名映射
TEST_COLUMN_MAP = {
    "title": ["title", "标题", "用例标题", "名称"],
    "module": ["module", "模块", "所属模块"],
    "priority": ["priority", "优先级"],
    "precondition": ["precondition", "前置条件", "前提条件"],
    "steps": ["steps", "步骤", "测试步骤"],
    "tags": ["tags", "标签"],
}

# 检查用例列名映射
CHECK_COLUMN_MAP = {
    "title": ["检查项目", "项目"],
    "module": ["类别", "分类"],
    "check_criteria": ["评价标准", "评价标準", "标准"],
    "check_category": ["类别", "分类"],
}


def _find_column(cols: list[str], candidates: list[str]) -> str | None:
    """在列名列表中查找匹配的列"""
    for c in candidates:
        c_lower = c.lower().strip()
        for col in cols:
            if col.lower().strip() == c_lower:
                return col
    return None


async def parse_import_file(file: UploadFile) -> dict:
    """解析上传的文件，自动识别类型，返回预览数据"""
    content = await file.read()
    errors = []
    valid_rows = []

    try:
        if file.filename and file.filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content), header=None)
    except Exception as e:
        return {"valid_rows": [], "errors": [{"row": 0, "field": "file", "message": f"文件解析失败: {str(e)}"}]}

    if df.empty:
        return {"valid_rows": [], "errors": [{"row": 0, "field": "file", "message": "文件中没有数据"}]}

    # 判断文件类型：检查是否有"检查项目"列
    columns = [str(c).strip() for c in df.iloc[0].tolist()]
    all_text = " ".join([str(v) for v in df.values.flatten() if pd.notna(v)])

    is_check = "检查项目" in all_text or "评价标准" in all_text or "评价标準" in all_text

    if is_check:
        return _parse_check_file(df, errors)
    else:
        return _parse_test_file(df, errors)


def _parse_test_file(df, errors) -> dict:
    """解析测试用例文件"""
    df.columns = [str(c).strip() for c in df.iloc[0].tolist()]
    df = df.iloc[1:]  # 去掉标题行

    columns = list(df.columns)
    mapping = {}
    for field, candidates in TEST_COLUMN_MAP.items():
        col = _find_column(columns, candidates)
        if col:
            mapping[col] = field

    if "title" not in mapping.values():
        return {"file_type": "test", "valid_rows": [], "errors": [{"row": 0, "field": "title", "message": "找不到标题/名称列"}]}

    inv_mapping = {v: k for k, v in mapping.items()}
    valid_rows = []

    for idx, row in df.iterrows():
        title = str(row.get(inv_mapping.get("title", ""), "")).strip()
        if not title or title == "nan":
            continue

        module = str(row.get(inv_mapping.get("module", ""), "")).strip()
        if not module or module == "nan":
            module = "默认模块"

        priority_raw = str(row.get(inv_mapping.get("priority", ""), "P2")).strip().upper()
        if priority_raw not in ("P0", "P1", "P2", "P3"):
            priority_raw = "P2"

        precondition = str(row.get(inv_mapping.get("precondition", ""), ""))
        if precondition == "nan":
            precondition = ""

        steps_raw = str(row.get(inv_mapping.get("steps", ""), ""))
        if steps_raw and steps_raw != "nan":
            steps = [{"seq": i + 1, "action": s.strip(), "expected": ""} for i, s in enumerate(steps_raw.split("\n")) if s.strip()]
        else:
            steps = []

        tags_raw = str(row.get(inv_mapping.get("tags", ""), ""))
        if tags_raw and tags_raw != "nan":
            tags = [t.strip() for t in tags_raw.replace("，", ",").split(",") if t.strip()]
        else:
            tags = []

        valid_rows.append({
            "title": title,
            "module": module,
            "priority": priority_raw,
            "status": "active",
            "precondition": precondition,
            "steps": steps,
            "tags": tags,
            "case_type": "test",
            "_row": int(idx) + 3,
        })

    return {"file_type": "test", "valid_rows": valid_rows, "errors": errors}


def _parse_check_file(df, errors) -> dict:
    """解析检查用例文件"""
    # 检查用例的列结构：序号 | 类别 | 检查项目 | 评价标准 | 频次 | 测试结果
    valid_rows = []
    all_text = "\n".join([str(v) for v in df.values.flatten() if pd.notna(v)])
    current_category = ""

    seq = 0
    for idx in range(df.shape[0]):
        row = df.iloc[idx]
        vals = [str(v).strip() for v in row.tolist() if pd.notna(v) and str(v).strip() != "nan"]

        if len(vals) < 2:
            continue

        # 跳过表头行
        if any(kw in " ".join(vals) for kw in ["检查项目", "评价标准", "测试结果", "车辆信息", "VIN"]):
            continue

        # 第一列可能是序号或类别名称
        first = vals[0]
        second = vals[1] if len(vals) > 1 else ""
        third = vals[2] if len(vals) > 2 else ""
        fourth = vals[3] if len(vals) > 3 else ""
        fifth = vals[4] if len(vals) > 4 else ""

        # 判断是类别标题还是检查项
        if first.isdigit() or (first.replace(".", "").isdigit() and len(vals) >= 3):
            # 是检查项：序号 | 检查项目 | 评价标准
            seq = int(float(first)) if first.replace(".", "").isdigit() else seq + 1
            if not all_text.split("\n")[0]:  # No specific category set yet
                pass

            # 查找真正的内容
            item_name = second if second and not second.isdigit() else third
            criteria = third if second and not second.isdigit() else fourth if len(vals) > 3 else ""

            if not item_name or item_name.isdigit():
                continue

            valid_rows.append({
                "title": item_name,
                "module": current_category or "检查",
                "check_category": current_category or "检查",
                "check_criteria": criteria,
                "priority": "P2",
                "status": "active",
                "case_type": "check",
                "precondition": "",
                "steps": [],
                "tags": [],
                "_row": int(idx) + 1,
            })
        else:
            # 可能是类别标题
            if first and not first.isdigit() and len(vals) <= 3:
                current_category = first

    return {"file_type": "check", "valid_rows": valid_rows, "errors": errors}
